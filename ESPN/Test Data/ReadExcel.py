import openpyxl

email = []
password = []


def read_excel():
    workbook = openpyxl.load_workbook("ESPN/TestData/TestData.xlsx")
    s = workbook.active
    rows = s.max_row
    cols = s.max_column
    for r in range(2, rows + 1):
        for c in range(1, cols + 1):
            if (c % 2 == 0):
                password.append(s.cell(row=r, column=c).value)

            else:
                email.append(s.cell(row=r, column=c).value)