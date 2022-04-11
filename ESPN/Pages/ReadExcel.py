import openpyxl


class ReadExcel:
    email = []
    password = []

    @staticmethod
    def read_excel():
        workbook = openpyxl.load_workbook("./ESPN/TestData/TestData.xlsx")
        s = workbook.active
        rows = s.max_row
        cols = s.max_column
        for r in range(2, rows + 1):
            for c in range(1, cols + 1):
                if (c % 2 == 0):
                    ReadExcel.password.append(s.cell(row=r, column=c).value)

                else:
                    ReadExcel.email.append(s.cell(row=r, column=c).value)
